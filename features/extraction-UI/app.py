"""
DocIntel Extraction Helper
---------------------------
A standalone Streamlit app that talks to an existing DocIntel FastAPI
backend to run repeatable, schema-based extraction over batches of
documents and collect the results into downloadable tables.

Auth: interactive sign-in only (POST /auth/login). Tokens live in
st.session_state for the duration of the browser session - nothing is
persisted to disk, and there is no env-var token fallback.

Run with:
    streamlit run app.py
"""
from datetime import datetime, timezone
from pathlib import Path

import html as html_lib
import io
import time

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from api_client import ApiError, DocIntelClient, extract_value
from batch_runner import run_batch
from config_store import get_base_url, get_timeout, get_max_parallel_uploads
from results_store import list_tables, load_table, save_table, delete_table, rename_table, to_excel_bytes
from schema_store import (
    list_schema_names, get_schema, save_schema, delete_schema,
    normalize_field, flatten_fields_for_table,
)

ALLOWED_EXTENSIONS = {".pdf", ".docx", ".csv", ".xlsx", ".txt", ".md", ".jpg", ".jpeg", ".png"}
FIELD_TYPES = ["string", "integer", "date", "list"]
EMPTY_FIELD = {"name": "", "type": "string", "description": "", "properties": None}

# import api_client
# print(api_client.__file__)
# print(hasattr(api_client.DocIntelClient, "list_agent_runs"))

st.set_page_config(page_title="DocIntel Extraction Helper", layout="wide")

# ----------------------------------------------------------------------
# Session state
# ----------------------------------------------------------------------
for key, default in [
    ("access_token", None),
    ("refresh_token", None),
    ("user_email", None),
    ("manual_fields", [dict(EMPTY_FIELD)]),
    ("last_run_results", None),       # list of {"filename": ..., "extracted": {...}} - nested, pre-flatten
    ("last_run_schema_name", None),
    ("nl_previewed_instruction", None),
    ("nl_preview_fields", []),
    ("agent_run_id", None),
    ("agent_run_status", None),        # last polled row: status/current_stage/pending_questions/result/error
    ("agent_answers_form", {}),
    ("chat_run_id", None),
    ("chat_send_error", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


def get_client() -> DocIntelClient:
    return DocIntelClient(get_base_url(), access_token=st.session_state.access_token or "", timeout=get_timeout())


def _clear_session():
    for k in ("access_token", "refresh_token", "user_email"):
        st.session_state[k] = None


def ensure_valid_session() -> bool:
    """Confirms the current token still works; tries one silent refresh if not."""
    client = get_client()
    try:
        client.me()
        return True
    except ApiError as e:
        if e.status_code == 401 and st.session_state.refresh_token:
            try:
                refreshed = client.refresh(st.session_state.refresh_token)
                st.session_state.access_token = refreshed["access_token"]
                st.session_state.refresh_token = refreshed.get("refresh_token", st.session_state.refresh_token)
                return True
            except ApiError:
                pass
        _clear_session()
        return False


def render_login():
    st.markdown(
        """
        <style>
        div[data-testid="stForm"] { max-width: 400px; margin: 0 auto; }
        </style>
        """,
        unsafe_allow_html=True,
    )
    _, mid, _ = st.columns([1, 1.2, 1])
    with mid:
        st.markdown("<div style='text-align:center'>📄</div>", unsafe_allow_html=True)
        st.markdown("<h2 style='text-align:center'>DocIntel Extraction Helper</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center;color:gray'>Sign in to continue</p>", unsafe_allow_html=True)
        with st.form("login_form"):
            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Sign in", type="primary", use_container_width=True)
        if submitted:
            if not email.strip() or not password:
                st.error("Enter both email and password.")
            else:
                try:
                    client = DocIntelClient(get_base_url(), timeout=get_timeout())
                    result = client.login(email.strip(), password)
                    st.session_state.access_token = result["access_token"]
                    st.session_state.refresh_token = result.get("refresh_token")
                    st.session_state.user_email = result.get("user", {}).get("email", email.strip())
                    st.rerun()
                except ApiError as e:
                    st.error(f"Sign in failed: {e}")


# ----------------------------------------------------------------------
# Recursive field editor - used by both "Manual fields" and the NL preview,
# so a list field can be expanded into its own nested sub-field rows.
# ----------------------------------------------------------------------
def render_field_editor(fields: list, key_prefix: str, show_include: bool = False, depth: int = 0):
    """
    Renders one row per field (name / type / description / delete), and for
    any field with type == "list", a nested, indented editor for its
    "properties" sub-fields. Mutates `fields` in place.
    """
    indent = "　" * depth  # ideographic space, gives a visible indent in Streamlit text
    to_delete = None
    for i, f in enumerate(fields):
        f.setdefault("type", "string")
        f.setdefault("properties", None)
        cols = st.columns([0.4, 1.6, 1, 3, 0.4]) if show_include else st.columns([1.6, 1, 3, 0.4])
        c = iter(cols)
        if show_include:
            f["include"] = next(c).checkbox("", value=f.get("include", True), key=f"{key_prefix}_inc_{i}")
        f["name"] = next(c).text_input(f"{indent}Field name", value=f["name"], key=f"{key_prefix}_name_{i}")
        f["type"] = next(c).selectbox(
            "Type", FIELD_TYPES, index=FIELD_TYPES.index(f["type"]) if f["type"] in FIELD_TYPES else 0,
            key=f"{key_prefix}_type_{i}",
        )
        f["description"] = next(c).text_input("Description", value=f["description"], key=f"{key_prefix}_desc_{i}")
        if next(c).button("✕", key=f"{key_prefix}_del_{i}"):
            to_delete = i

        if f["type"] == "list":
            st.caption(f"{indent}↳ sub-fields of **{f['name'] or '(unnamed list)'}**")
            if not f.get("properties"):
                f["properties"] = [dict(EMPTY_FIELD)]
            render_field_editor(f["properties"], f"{key_prefix}_sub_{i}", show_include=False, depth=depth + 1)
            if st.button(f"{indent}+ Add sub-field", key=f"{key_prefix}_addsub_{i}"):
                f["properties"].append(dict(EMPTY_FIELD))
                st.rerun()
        else:
            f["properties"] = None

    if to_delete is not None and len(fields) > 1:
        fields.pop(to_delete)
        st.rerun()


def clean_fields_for_save(fields: list, require_include: bool = False) -> list:
    """Drops unnamed rows and (for NL preview) unchecked ones, recursively."""
    out = []
    for f in fields:
        if not f["name"].strip():
            continue
        if require_include and not f.get("include", True):
            continue
        cleaned = normalize_field(f)
        if cleaned["type"] == "list" and f.get("properties"):
            cleaned["properties"] = clean_fields_for_save(f["properties"], require_include=False)
        out.append(cleaned)
    return out


# ----------------------------------------------------------------------
# Nested results rendering - one document per expander: scalar fields as
# a key/value summary, each list field as its own sub-table underneath.
# ----------------------------------------------------------------------
def build_document_rows(extracted: dict):
    """
    Shapes one document's extracted data into the Excel-style layout:
    - scalar fields -> one value, only on row 0 (blank on later rows -> merge target)
    - each list field -> its own sub-columns ("field.subfield"), one row per
      item; rows beyond that field's own item count are blank
    Row count for the document = max(len of any list field, 1).
    Returns (columns, rows, scalar_col_names).
    """
    scalar, list_fields = {}, {}
    for field_name, raw_value in extracted.items():
        value = extract_value(raw_value)
        if isinstance(value, list):
            list_fields[field_name] = value
        else:
            scalar[field_name] = value

    list_columns = {}
    for fname, items in list_fields.items():
        cols, seen = [], set()
        for item in items:
            if isinstance(item, dict):
                for k in item.keys():
                    if k not in seen:
                        seen.add(k)
                        cols.append(k)
        list_columns[fname] = cols or ["value"]

    max_len = max([len(v) for v in list_fields.values()] + [1])
    rows = []
    for i in range(max_len):
        row = dict(scalar) if i == 0 else {k: None for k in scalar}
        for fname, items in list_fields.items():
            cols = list_columns[fname]
            item = items[i] if i < len(items) else None
            if item is None:
                for c in cols:
                    row[f"{fname}.{c}"] = None
            elif isinstance(item, dict):
                for c in cols:
                    row[f"{fname}.{c}"] = item.get(c)
            else:
                row[f"{fname}.value"] = item
        rows.append(row)

    scalar_col_names = list(scalar.keys())
    list_col_names = [f"{fname}.{c}" for fname, cols in list_columns.items() for c in cols]
    return scalar_col_names + list_col_names, rows, scalar_col_names


def _shaped_docs(results: list):
    """Runs build_document_rows for every successful result and unions columns/scalar-col-set across docs."""
    docs, columns, scalar_cols_all = [], [], set()
    for r in results:
        if r.get("error"):
            continue
        cols, rows, scalar_cols = build_document_rows(r.get("extracted", {}))
        docs.append((r.get("filename", "unknown"), r.get("document_id", ""), rows, scalar_cols))
        for c in cols:
            if c not in columns:
                columns.append(c)
        scalar_cols_all.update(scalar_cols)
    return docs, columns, scalar_cols_all


def render_merged_results_table(results: list):
    """Renders the whole run as one Excel-style HTML table: one column per field, filename + scalar fields
    row-span-merged down each document's block of rows, list fields filling their own rows underneath."""
    errored = [r for r in results if r.get("error")]
    docs, columns, scalar_cols_all = _shaped_docs(results)

    if not docs:
        st.info("No successful extractions to display.")
    else:
        header_cols = ["filename", "document_id"] + columns
        parts = [
            "<div style='overflow-x:auto'><table style='border-collapse:collapse;width:100%;font-size:0.85rem'>",
            "<thead><tr>" + "".join(
                f"<th style='padding:6px 8px;background:#262730;color:white;text-align:left;position:sticky;top:0'>{html_lib.escape(c)}</th>"
                for c in header_cols
            ) + "</tr></thead><tbody>",
        ]
        for filename, document_id, rows, scalar_cols in docs:
            n = len(rows)
            for i, row in enumerate(rows):
                parts.append("<tr>")
                if i == 0:
                    parts.append(
                        f"<td rowspan='{n}' style='padding:6px 8px;vertical-align:top;font-weight:600;border:1px solid #444'>{html_lib.escape(filename)}</td>"
                    )
                    parts.append(
                        f"<td rowspan='{n}' style='padding:6px 8px;vertical-align:top;color:#888;font-size:0.75rem;border:1px solid #444'>{html_lib.escape(document_id)}</td>"
                    )
                for c in columns:
                    if c in scalar_cols:
                        if i == 0:
                            val = row.get(c)
                            parts.append(
                                f"<td rowspan='{n}' style='padding:6px 8px;vertical-align:top;border:1px solid #444'>{'' if val is None else html_lib.escape(str(val))}</td>"
                            )
                        # else: cell is covered by the rowspan above - emit nothing
                    else:
                        val = row.get(c)
                        parts.append(
                            f"<td style='padding:6px 8px;border:1px solid #444'>{'' if val is None else html_lib.escape(str(val))}</td>"
                        )
                parts.append("</tr>")
        parts.append("</tbody></table></div>")
        st.markdown("".join(parts), unsafe_allow_html=True)

    if errored:
        with st.expander(f"⚠️ {len(errored)} file(s) had errors"):
            for r in errored:
                st.write(f"- {r.get('filename', 'unknown')}: {r['error']}")


def to_excel_bytes_merged(results: list, sheet_name: str = "Extraction Results") -> bytes:
    """Same layout as render_merged_results_table, but as a real .xlsx with actual merged cells
    (openpyxl merge_cells) instead of HTML rowspan - opens correctly in Excel/Sheets."""
    docs, columns, scalar_cols_all = _shaped_docs(results)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"

    header = ["filename", "document_id"] + columns
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    current_row = 2
    for filename, document_id, rows, scalar_cols in docs:
        n = len(rows)
        start_row = current_row
        for i, row in enumerate(rows):
            ws.append([filename if i == 0 else None, document_id if i == 0 else None] + [row.get(c) for c in columns])
        end_row = start_row + n - 1
        if n > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            for idx, c in enumerate(columns, start=3):
                if c in scalar_cols:
                    ws.merge_cells(start_row=start_row, start_column=idx, end_row=end_row, end_column=idx)
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
            ws.cell(row=r, column=2).alignment = Alignment(vertical="top")
        current_row = end_row + 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------
# Agents - polling + result rendering
# ----------------------------------------------------------------------
AGENT_TERMINAL_STATUSES = {"needs_input", "completed", "failed"}


def poll_agent_run_until_terminal(client: DocIntelClient, run_id: str, poll_interval: float = 2.0, timeout: float = 600.0) -> dict:
    """
    Blocks (with a spinner) polling GET /agents/runs/{run_id} until status
    reaches needs_input/completed/failed, or `timeout` seconds elapse.
    Matches the existing blocking-with-spinner pattern used for batch
    extraction runs - agent runs are similarly long (multiple LLM calls),
    so this is consistent UX rather than a background-refresh model.
    """
    status_area = st.empty()
    start = time.time()
    run = client.get_agent_run(run_id)
    while run.get("status") not in AGENT_TERMINAL_STATUSES:
        elapsed = time.time() - start
        if elapsed > timeout:
            run["status"] = "failed"
            run["error"] = f"Timed out waiting for the agent after {int(timeout)}s (last stage: {run.get('current_stage')})."
            break
        status_area.write(f"⏳ {run.get('status', 'pending')} — stage: {run.get('current_stage') or 'starting'} ({int(elapsed)}s)")
        time.sleep(poll_interval)
        run = client.get_agent_run(run_id)
    status_area.empty()
    return run


def render_agent_result(result: dict):
    """Renders the fixed agent output contract: summary (prose), findings (bullets), data (tables/csv/json)."""
    if result.get("summary"):
        st.write(result["summary"])

    findings = result.get("findings") or []
    if findings:
        st.markdown("\n".join(f"- {f}" for f in findings))

    for item in result.get("data") or []:
        label = item.get("label") or item.get("type", "Data")
        item_type = item.get("type")
        value = item.get("value")
        st.caption(f"**{label}**")
        if item_type == "table" and isinstance(value, list):
            st.dataframe(pd.DataFrame(value), use_container_width=True)
        elif item_type == "csv":
            st.code(value if isinstance(value, str) else str(value), language="text")
        elif item_type == "json":
            st.json(value)
        else:
            st.write(value)


def _sanitize_records_for_json(records: list[dict]) -> list[dict]:
    """
    pandas.DataFrame.where(pd.notnull(df), None) does NOT reliably turn NaN
    into None for float64 columns - pandas converts None back to NaN when
    writing into a numeric column, so the NaN survives to_dict(). httpx's
    JSON encoder calls json.dumps(..., allow_nan=False), which raises
    ValueError on any leftover NaN/Infinity - so every value needs an
    explicit, type-level check here rather than relying on pandas' fill.
    """
    import math

    def _clean(v):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if v is pd.NaT:
            return None
        return v

    return [{k: _clean(v) for k, v in row.items()} for row in records]


def flatten_results_for_export(results: list) -> pd.DataFrame:
    """
    Denormalized flat export: scalar fields repeat/blank per build_document_rows'
    row shaping, and EVERY list field gets its own sub-columns - reuses the
    exact same row-shaping logic as the merged view/Excel export (build_document_rows)
    so the saved CSV table is never missing a list field just because it wasn't
    the first one encountered. (Previously this had its own simplified
    single-list-field implementation that silently dropped every other list
    field and under-counted rows - fixed by reusing build_document_rows here.)
    """
    rows = []
    for r in results:
        filename = r.get("filename")
        document_id = r.get("document_id")
        if r.get("error"):
            rows.append({"filename": filename, "document_id": document_id, "error": r["error"]})
            continue
        _cols, doc_rows, _scalar_cols = build_document_rows(r.get("extracted", {}))
        for doc_row in doc_rows:
            rows.append({"filename": filename, "document_id": document_id, **doc_row})
    return pd.DataFrame(rows)


# ----------------------------------------------------------------------
# Auth gate
# ----------------------------------------------------------------------
if not st.session_state.access_token:
    render_login()
    st.stop()

if not ensure_valid_session():
    st.warning("Your session expired. Please sign in again.")
    render_login()
    st.stop()

# ----------------------------------------------------------------------
# Top bar - account controls (no sidebar)
# ----------------------------------------------------------------------
top_l, top_r = st.columns([4, 1])
with top_l:
    st.title("📄 DocIntel Extraction Helper")
    st.caption("Define an extraction schema once, run it over batches of documents, and build up a results table.")
with top_r:
    st.write("")
    st.caption(f"Signed in as **{st.session_state.user_email}**")
    if st.button("Sign out", use_container_width=True):
        try:
            get_client().logout()
        except ApiError:
            pass  # best-effort - clear local session regardless
        _clear_session()
        st.rerun()

with st.expander("🔒 Change password"):
    with st.form("change_password_form"):
        old_password = st.text_input("Current password", type="password")
        new_password = st.text_input("New password", type="password")
        confirm_password = st.text_input("Confirm new password", type="password")
        change_submitted = st.form_submit_button("Change password")
    if change_submitted:
        if not old_password or not new_password:
            st.error("Fill in both current and new password.")
        elif new_password != confirm_password:
            st.error("New password and confirmation don't match.")
        elif len(new_password) < 8:
            st.error("New password must be at least 8 characters.")
        else:
            try:
                verify_client = DocIntelClient(get_base_url(), timeout=get_timeout())
                verify_result = verify_client.login(st.session_state.user_email, old_password)
                verify_client.reset_password(verify_result["access_token"], new_password)
                st.session_state.access_token = verify_result["access_token"]
                st.session_state.refresh_token = verify_result.get("refresh_token")
                st.success("Password changed.")
            except ApiError as e:
                if e.status_code == 401:
                    st.error("Current password is incorrect.")
                else:
                    st.error(f"Could not change password: {e}")

TAB_NAMES = [
            "1️⃣ Build / Save Schema", "2️⃣ Run Extraction", "3️⃣ Saved Tables",
            "4️⃣ Agents", "5️⃣ Past Runs", "6️⃣ Chat",
        ]
if st.session_state.get("active_tab") not in TAB_NAMES:
    st.session_state.active_tab = TAB_NAMES[0]
st.markdown(
    """
    <style>
    div[role="radiogroup"] { gap: 24px; border-bottom: 2px solid #444; padding-bottom: 8px; }
    div[role="radiogroup"] label { font-size: 1.05rem; font-weight: 500; }
    </style>
    """,
    unsafe_allow_html=True,
)
st.session_state.active_tab = st.radio(
    "Navigate", TAB_NAMES, horizontal=True,
    index=TAB_NAMES.index(st.session_state.active_tab),
    label_visibility="collapsed",
)
active_tab = st.session_state.active_tab

# ----------------------------------------------------------------------
# TAB 1 - Schema builder
# ----------------------------------------------------------------------
if active_tab == TAB_NAMES[0]:
    left, right = st.columns([1, 1])

    with left:
        st.subheader("Saved schemas")
        names = list_schema_names()
        if names:
            for n in names:
                s = get_schema(n)
                with st.expander(f"**{n}**"):
                    st.table(pd.DataFrame(flatten_fields_for_table(s.get("fields", []))))
                    if st.button("Delete schema", key=f"del_schema_{n}"):
                        delete_schema(n)
                        st.rerun()
        else:
            st.info("No schemas saved yet — build one on the right.")

        st.subheader("Load from engine templates")
        st.caption("Pulls pre-built schemas from GET /templates so you don't have to redefine common ones by hand.")
        if st.button("🔄 Fetch templates"):
            try:
                with st.spinner("Fetching templates..."):
                    templates = get_client().get_templates()
                st.session_state["_fetched_templates"] = templates
            except ApiError as e:
                st.error(str(e))

        templates = st.session_state.get("_fetched_templates")
        if templates:
            # ASSUMPTION: each template item has a name (name/template_name)
            # and a fields list (name/type/description/properties, same
            # shape as everywhere else in this app). Adjust the key lookups
            # below if the real /templates response differs.
            for i, tpl in enumerate(templates):
                tpl_name = tpl.get("name") or tpl.get("template_name") or f"template_{i}"
                tpl_fields = tpl.get("fields") or tpl.get("schema", {}).get("fields", [])
                with st.expander(f"📋 {tpl_name}"):
                    if tpl.get("description"):
                        st.caption(tpl["description"])
                    if tpl_fields:
                        st.table(pd.DataFrame(flatten_fields_for_table(normalize_fields(tpl_fields))))
                    else:
                        st.json(tpl)  # unrecognized shape - show raw so it can still be inspected
                    save_as = st.text_input("Save as schema name", value=tpl_name, key=f"tpl_save_name_{i}")
                    if st.button("💾 Save this template as a schema", key=f"tpl_save_{i}", disabled=not tpl_fields):
                        save_schema(save_as.strip() or tpl_name, {"mode": "fields", "fields": normalize_fields(tpl_fields)})
                        st.success(f"Saved '{save_as or tpl_name}'.")
                        st.rerun()
        elif "_fetched_templates" in st.session_state:
            st.info("No templates returned by the engine.")

    with right:
        st.subheader("Create a new schema")
        schema_name = st.text_input("Schema name", placeholder="e.g. resume_extraction")
        mode = st.radio("How do you want to define fields?", ["Manual fields", "From NL instruction (preview)"])

        # ---------------- Manual fields ----------------
        if mode == "Manual fields":
            st.caption("Define the fields you want extracted. Set a field's type to 'list' to add nested sub-fields.")
            render_field_editor(st.session_state.manual_fields, "manual")
            if st.button("+ Add field"):
                st.session_state.manual_fields.append(dict(EMPTY_FIELD))
                st.rerun()

            if st.button("💾 Save schema", type="primary"):
                fields = clean_fields_for_save(st.session_state.manual_fields)
                if not schema_name.strip():
                    st.error("Give the schema a name first.")
                elif not fields:
                    st.error("Add at least one field with a name.")
                else:
                    save_schema(schema_name.strip(), {"mode": "fields", "fields": fields})
                    st.success(f"Saved schema '{schema_name}'")
                    st.session_state.manual_fields = [dict(EMPTY_FIELD)]
                    st.rerun()

        # ---------------- NL instruction -> schema preview ----------------
        else:
            st.caption(
                "Describe in plain English what to extract, then preview the schema it derives — "
                "no document needed. Review/edit the fields (including nested sub-fields for list "
                "fields), then save."
            )
            instruction = st.text_area(
                "Instruction",
                placeholder="Extract candidate name, education (institute, course, branch, start/end year), "
                            "and work experience (company, location, start/end date, designation).",
                height=120,
            )

            if st.button("🔍 Preview schema"):
                if not instruction.strip():
                    st.error("Write an instruction first.")
                else:
                    try:
                        client = get_client()
                        with st.spinner("Generating schema preview..."):
                            preview_fields = client.preview_nl_schema(instruction.strip())
                        st.session_state.nl_previewed_instruction = instruction.strip()
                        st.session_state.nl_preview_fields = [
                            {**normalize_field(f), "include": True} for f in preview_fields
                        ]
                        st.rerun()
                    except ApiError as e:
                        st.error(str(e))

            previewed = st.session_state.nl_previewed_instruction
            instruction_changed = previewed is not None and previewed != instruction.strip()

            if previewed and not instruction_changed:
                if st.session_state.nl_preview_fields:
                    st.success("Preview generated — review and edit the fields below:")
                    render_field_editor(st.session_state.nl_preview_fields, "nl", show_include=True)
                else:
                    st.warning("Preview came back with no fields — try rewording the instruction.")
            elif instruction_changed:
                st.warning("Instruction changed since your last preview — preview again before saving.")
            else:
                st.info("Preview the schema above before you can save it.")

            can_save = bool(previewed) and not instruction_changed and bool(st.session_state.nl_preview_fields)
            if st.button("💾 Save schema", type="primary", key="save_nl_schema", disabled=not can_save):
                fields = clean_fields_for_save(st.session_state.nl_preview_fields, require_include=True)
                if not schema_name.strip():
                    st.error("Give the schema a name first.")
                elif not fields:
                    st.error("At least one previewed field needs to stay included with a name.")
                else:
                    save_schema(schema_name.strip(), {"mode": "fields", "fields": fields})
                    st.success(f"Saved schema '{schema_name}' ({len(fields)} field(s))")
                    st.session_state.nl_previewed_instruction = None
                    st.session_state.nl_preview_fields = []
                    st.rerun()

# ----------------------------------------------------------------------
# TAB 2 - Run extraction (parallel, pipelined upload -> extract per file)
# ----------------------------------------------------------------------
if active_tab == TAB_NAMES[1]:
    schema_names = list_schema_names()
    if not schema_names:
        st.warning("No saved schemas yet. Create one in the 'Build / Save Schema' tab first.")
    else:
        chosen_schema_name = st.selectbox("Schema to use", schema_names)
        schema = get_schema(chosen_schema_name)

        st.subheader("Select documents")
        source = st.radio("Document source", ["Upload files", "Local folder path"], horizontal=True)

        files_to_process = []  # list of (filename, bytes)

        if source == "Upload files":
            uploaded = st.file_uploader("Choose documents", accept_multiple_files=True)
            if uploaded:
                files_to_process = [(f.name, f.getvalue()) for f in uploaded]
        else:
            folder = st.text_input("Folder path (on the machine running this app)")
            if folder:
                p = Path(folder).expanduser()
                if p.is_dir():
                    found = [f for f in sorted(p.iterdir()) if f.suffix.lower() in ALLOWED_EXTENSIONS and f.is_file()]
                    st.write(f"Found {len(found)} matching file(s).")
                    if found:
                        st.write([f.name for f in found])
                        files_to_process = [(f.name, f.read_bytes()) for f in found]
                else:
                    st.error("That path doesn't exist or isn't a folder.")

        st.subheader("Destination table")
        st.caption(
            "Saved as a flattened Excel/CSV table (list fields exploded into rows). "
            "The nested view below the run button is always the full-fidelity source of truth."
        )
        existing_tables = list_tables()
        dest_mode = st.radio("Save results to", ["New table", "Append to existing table"], horizontal=True)
        if dest_mode == "New table":
            default_name = f"{chosen_schema_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}"
            table_name = st.text_input("New table name", value=default_name)
        else:
            if existing_tables:
                table_name = st.selectbox("Existing table", existing_tables)
            else:
                st.info("No existing tables yet — a new one will be created instead.")
                table_name = f"{chosen_schema_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}"
                dest_mode = "New table"

        max_parallel = get_max_parallel_uploads()

        run_disabled = not files_to_process
        if st.button("▶️ Run extraction", type="primary", disabled=run_disabled):
            progress = st.progress(0.0)
            status = st.empty()

            def _on_progress(completed, total, fname, ok):
                progress.progress(completed / total)
                mark = "✅" if ok else "❌"
                status.write(f"{mark} {fname} ({completed}/{total})")

            with st.spinner(f"Processing {len(files_to_process)} file(s)...."):
                rows, errors = run_batch(
                    base_url=get_base_url(),
                    access_token=st.session_state.access_token,
                    timeout=get_timeout(),
                    files=files_to_process,
                    fields=schema["fields"],
                    max_parallel=max_parallel,
                    schema_name=chosen_schema_name,
                    progress_cb=_on_progress,
                )

            status.write("Done.")
            st.session_state.last_run_results = rows  # expected: list of {"filename","extracted","error"?}
            st.session_state.last_run_schema_name = chosen_schema_name

            if errors:
                st.warning(f"{len(errors)} file(s) had errors — see details below.")
                with st.expander("Error details"):
                    for e in errors:
                        st.write("- " + e)

            flat_df = flatten_results_for_export(rows)
            saved_df = save_table(table_name, flat_df, mode="append" if dest_mode == "Append to existing table" else "new")
            st.success(f"Saved {len(flat_df)} row(s) to table '{table_name}' ({len(saved_df)} total rows).")

        if st.session_state.last_run_results is not None:
            st.subheader("Latest run results")
            render_merged_results_table(st.session_state.last_run_results)

            st.download_button(
                "⬇️ Download this run as Excel (merged cells)",
                data=to_excel_bytes_merged(st.session_state.last_run_results, sheet_name=st.session_state.last_run_schema_name or "Results"),
                file_name=f"{st.session_state.last_run_schema_name or 'results'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ----------------------------------------------------------------------
# TAB 3 - Saved tables (flattened - see caption above in Tab 2)
# ----------------------------------------------------------------------
if active_tab == TAB_NAMES[2]:
    tables = list_tables()
    if not tables:
        st.info("No saved tables yet. Run an extraction to create one.")
    else:
        selected_table = st.selectbox("Table", tables)
        df = load_table(selected_table)
        st.write(f"{len(df)} row(s)")
        st.dataframe(df, use_container_width=True)

        with st.expander("✏️ Rename table"):
            new_name = st.text_input("New name", value=selected_table, key=f"rename_{selected_table}")
            if st.button("Rename", key=f"rename_btn_{selected_table}"):
                try:
                    renamed_to = rename_table(selected_table, new_name)
                    st.success(f"Renamed to '{renamed_to}'.")
                    st.rerun()
                except (FileNotFoundError, ValueError) as e:
                    st.error(str(e))

        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                "⬇️ Download as Excel",
                data=to_excel_bytes(df, sheet_name=selected_table),
                file_name=f"{selected_table}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with c2:
            if st.button("🗑️ Delete table"):
                delete_table(selected_table)
                st.rerun()


# ----------------------------------------------------------------------
# TAB 4 - Agents (invoke, poll, answer clarifying questions, view result)
# ----------------------------------------------------------------------
if active_tab == TAB_NAMES[3]:
    st.subheader("Run an agent over a batch of candidates")

    tables = list_tables()
    if not tables:
        st.info("No saved tables yet. Run an extraction first (Tab 2) - agents work over an extracted CSV table.")
    else:
        source_table = st.selectbox("Source table (candidates)", tables, key="agent_source_table")
        source_df = load_table(source_table)

        if "document_id" not in source_df.columns:
            st.warning(
                "This table has no 'document_id' column - it was likely saved before document_id was "
                "added to the CSV export. Re-run extraction on these documents to get a table agents can use."
            )
        elif source_df.empty:
            st.info("This table has no rows.")
        else:
            st.caption(f"{len(source_df)} row(s) in '{source_table}'.")
            select_all = st.checkbox("Include all rows", value=True)
            if select_all:
                subset_df = source_df
            else:
                display_col = "filename" if "filename" in source_df.columns else source_df.columns[0]
                chosen = st.multiselect("Candidates to include", source_df[display_col].tolist())
                subset_df = source_df[source_df[display_col].isin(chosen)]

            valid_subset = subset_df[subset_df["document_id"].notna() & (subset_df["document_id"] != "")]
            if len(valid_subset) < len(subset_df):
                st.caption(f"⚠️ Skipping {len(subset_df) - len(valid_subset)} row(s) with no document_id.")

            try:
                available_agents = get_client().list_agents()
            except ApiError:
                available_agents = ["cv_processor"]  # fallback if the endpoint isn't reachable yet
            agent_name = st.selectbox("Agent", available_agents or ["cv_processor"])

            with st.form("agent_invoke_form"):
                run_name = st.text_input(
                    "Name this run (optional)",
                    placeholder="e.g. Backend Engineer - July batch",
                )
                task_text = st.text_area(
                    "Task",
                    placeholder=(
                        "Identify the best 5 candidates for this JD. Main skills to evaluate: Python (must-have), "
                        "Docker (nice to have). Education is important - IIT/NIT preferred. ..."
                    ),
                    height=140,
                )
                invoke_disabled = valid_subset.empty or st.session_state.agent_run_id is not None
                invoke_submitted = st.form_submit_button("▶️ Run agent", type="primary", disabled=invoke_disabled)

            if invoke_submitted:
                if not task_text.strip():
                    st.error("Describe the task first.")
                else:
                    document_ids = valid_subset["document_id"].drop_duplicates().tolist()
                    csv_data = _sanitize_records_for_json(valid_subset.to_dict("records"))
                    try:
                        client = get_client()
                        invoke_result = client.invoke_agent(
                                agent_name, task_text.strip(), document_ids, csv_data,
                                name=run_name.strip() or None,
                            )
                        st.session_state.agent_run_id = invoke_result["run_id"]
                        with st.spinner("Agent running..."):
                            run = poll_agent_run_until_terminal(client, invoke_result["run_id"])
                        st.session_state.agent_run_status = run
                    except ApiError as e:
                        st.error(f"Could not start agent: {e}")

    # ---- Render whatever the last known run status is ----
    run = st.session_state.agent_run_status
    if run:
        status = run.get("status")
        st.divider()

        if status == "needs_input":
            questions = run.get("pending_questions") or []
            form_placeholder = st.empty()
            with form_placeholder.container():
                st.info("The agent needs a bit more information before continuing.")
                with st.form("agent_answers_form"):
                    answers = {}
                    for q in questions:
                        key, question, q_type, options = q["key"], q["question"], q.get("type", "text"), q.get("options") or []
                        if q_type == "select" and options:
                            answers[key] = st.selectbox(question, options, key=f"agent_q_{key}")
                        else:
                            answers[key] = st.text_input(question, key=f"agent_q_{key}")
                    submitted = st.form_submit_button("Submit answers and continue")

            if submitted:
                form_placeholder.empty()  # remove the form/questions UI immediately, before the blocking resume+poll below
                try:
                    client = get_client()
                    client.resume_agent_run(st.session_state.agent_run_id, answers)
                    with st.spinner("Agent resuming..."):
                        run = poll_agent_run_until_terminal(client, st.session_state.agent_run_id)
                    st.session_state.agent_run_status = run
                    st.rerun()
                except ApiError as e:
                    st.error(f"Could not resume agent: {e}")

        elif status == "completed":
            st.success("Agent finished.")
            render_agent_result(run.get("result") or {})
            c1, c2 = st.columns(2)
            with c1:
                if st.button("💬 Chat about this run"):
                    st.session_state.chat_run_id = run.get("id") or st.session_state.agent_run_id
                    st.session_state.active_tab = TAB_NAMES[5]
                    st.rerun()
            with c2:
                if st.button("Start a new agent run"):
                    st.session_state.agent_run_id = None
                    st.session_state.agent_run_status = None
                    st.rerun()

        elif status == "failed":
            st.error(f"Agent run failed: {run.get('error', 'unknown error')}")
            if st.button("Start a new agent run", key="retry_failed"):
                st.session_state.agent_run_id = None
                st.session_state.agent_run_status = None
                st.rerun()


# ----------------------------------------------------------------------
# TAB 5 - Past agent runs
# ----------------------------------------------------------------------
if active_tab == TAB_NAMES[4]:
    st.subheader("Past agent runs")
    try:
        client = get_client()
        runs = client.list_agent_runs()
    except ApiError as e:
        runs = []
        st.error(f"Could not load past runs: {e}")

    if not runs:
        st.info("No agent runs yet.")
    else:
        status_icon = {"completed": "✅", "failed": "❌", "needs_input": "⏸️", "running": "⏳", "pending": "⏳"}
        for run_summary in runs:
            icon = status_icon.get(run_summary.get("status"), "•")
            display_name = run_summary.get("name") or run_summary["id"]
            label = f"{icon} {display_name} — {run_summary.get('agent_name')} — {run_summary.get('created_at', '')[:19]} — {run_summary.get('status')}"
            with st.expander(label):
                st.caption(run_summary.get("task", ""))
                b1, b2 = st.columns(2)
                with b1:
                    if st.button("View details", key=f"view_run_{run_summary['id']}"):
                        try:
                            full_run = get_client().get_agent_run(run_summary["id"])
                            st.session_state[f"_past_run_detail_{run_summary['id']}"] = full_run
                        except ApiError as e:
                            st.error(f"Could not load run: {e}")
                with b2:
                    if run_summary.get("status") == "completed":
                        if st.button("💬 Chat", key=f"chat_run_{run_summary['id']}"):
                            st.session_state.chat_run_id = run_summary["id"]
                            st.session_state.active_tab = TAB_NAMES[5]
                            st.rerun()

                detail = st.session_state.get(f"_past_run_detail_{run_summary['id']}")
                if detail:
                    if detail.get("status") == "completed":
                        render_agent_result(detail.get("result") or {})
                    elif detail.get("status") == "failed":
                        st.error(f"Failed: {detail.get('error', 'unknown error')}")
                    elif detail.get("status") == "needs_input":
                        st.warning("This run is waiting on clarifying questions — answer it from the Agents tab.")
                    else:
                        st.write(f"Status: {detail.get('status')} — stage: {detail.get('current_stage')}")
                        
if active_tab == TAB_NAMES[5]:
    st.subheader("Chat about an agent run")
 
    try:
        completed_runs = get_client().list_agent_runs(status="completed")
    except ApiError as e:
        completed_runs = []
        st.error(f"Could not load completed runs: {e}")
 
    if not completed_runs:
        st.info("No completed agent runs yet — chat is only available once a run finishes.")
    else:
        def _run_label(r):
            display_name = r.get("name") or r["id"]
            return f"{display_name} — {r.get('agent_name', '')} ({r.get('created_at', '')[:19]})"
 
        run_ids = [r["id"] for r in completed_runs]
        run_labels = {r["id"]: _run_label(r) for r in completed_runs}
 
        default_id = st.session_state.chat_run_id if st.session_state.chat_run_id in run_ids else run_ids[0]
        selected_run_id = st.selectbox(
            "Agent run",
            run_ids,
            index=run_ids.index(default_id),
            format_func=lambda rid: run_labels[rid],
            key="chat_run_selector",
        )
        st.session_state.chat_run_id = selected_run_id
 
        st.divider()
 
        try:
            history = get_client().get_agent_chat_history(selected_run_id)
        except ApiError as e:
            history = []
            st.error(f"Could not load chat history: {e}")
 
        for msg in history:
            with st.chat_message(msg["role"]):
                st.write(msg["content"])
 
        
        if st.session_state.chat_send_error:
            st.error(st.session_state.chat_send_error)
 
        user_msg = st.chat_input("Ask about this run's result...")
        if user_msg:
            try:
                with st.spinner("Thinking..."):
                    get_client().send_chat_message(selected_run_id, user_msg)
                st.session_state.chat_send_error = None
            except ApiError as e:
                # NOTE: st.chat_input always clears on rerun (no way to
                # restore the typed text), so a failed send does mean
                # re-typing the message - the error is at least now visible
                # instead of flashing and disappearing.
                st.session_state.chat_send_error = f"Could not send message: {e}"
            st.rerun()